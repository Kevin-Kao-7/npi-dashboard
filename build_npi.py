"""
从 Excel 生成 NPI Dashboard 数据并嵌入 HTML。
用法: python build_npi.py
"""

import json
import re
import datetime
import hashlib
import openpyxl

# ── 路径配置 ──
EXCEL_PATH = r"C:\Users\msipm\Desktop\work\Spec總表.xlsx"
HTML_PATH  = r"c:\Users\msipm\WorkBuddy\20260422080636\npi_dashboard.html"
JSON_PATH  = r"c:\Users\msipm\WorkBuddy\20260422080636\npi_data.json"

# ── 密码配置 ──
DASHBOARD_PASSWORD = "msiyjb"

# Excel Schedule sheet 列映射 (1-indexed, Row 1=表头)
COL = {
    "model": 1, "mkt": 2, "series": 3, "segment": 4,
    "cpu": 5, "gpu": 6, "npm": 7, "spm": 8,
    "stage": 9,
    "id_frozen": 10,
    "Kickoff": 11, "DVT-start": 12, "EVT-start": 13,
    "MVT-start": 14, "BTO ready": 15, "ATS-start": 16,
    "MP": 17,
    "status": 18, "highlight": 19,
}

# 日期列名（按顺序）
DATE_KEYS = ["Kickoff", "DVT-start", "EVT-start", "MVT-start", "BTO ready", "ATS-start", "MP"]


def fmt_date(val):
    """把 datetime / 字符串 / na 转成 YYYY/MM/DD 格式字符串，无效返回 None"""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y/%m/%d")
    if isinstance(val, str):
        val = val.strip()
        if not val or val.lower() == "na" or val.lower() == "n/a":
            return None
        # 尝试解析常见日期格式
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%e", "%m/%d/%Y"):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                return dt.strftime("%Y/%m/%d")
            except ValueError:
                continue
        return val
    return None


def read_excel():
    """读取 Excel Schedule sheet 并返回 records 列表"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Schedule"]

    records = []
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row_idx, COL["model"]).value
        if not model or not str(model).strip():
            continue

        stage = ws.cell(row_idx, COL["stage"]).value
        if not stage or not str(stage).strip():
            continue

        # 构造 dates 对象
        dates = {}
        mp_val = None
        for dk in DATE_KEYS:
            v = fmt_date(ws.cell(row_idx, COL[dk]).value)
            if v is not None:
                dates[dk] = v
                if dk == "MP":
                    mp_val = v

        # 如果没有日期则跳过
        if not dates:
            continue

        record = {
            "model": str(model).strip(),
            "mkt": str(ws.cell(row_idx, COL["mkt"]).value or "").strip(),
            "cpu": str(ws.cell(row_idx, COL["cpu"]).value or "").strip(),
            "gpu": str(ws.cell(row_idx, COL["gpu"]).value or "").strip(),
            "npm": str(ws.cell(row_idx, COL["npm"]).value or "").strip(),
            "stage": str(stage).strip(),
            "status": str(ws.cell(row_idx, COL["status"]).value or "").strip(),
            "highlight": str(ws.cell(row_idx, COL["highlight"]).value or "").strip(),
            "mp_sort": mp_val or "",
            "orig_idx": len(records),
            "dates": dates,
        }
        records.append(record)

    wb.close()
    return records


def inject_html(records):
    """把数据写入 HTML 的 const DATA = { ... } 部分"""
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    # 把 records 序列化为 JSON
    data_obj = {"records": records, "buildTime": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")}
    json_str = json.dumps(data_obj, ensure_ascii=False, indent=2)

    # 替换 const DATA = { ... }; 块
    pattern = r"const DATA = \{[\s\S]*?\n\};"
    replacement = f"const DATA = {json_str};"

    new_html, count = re.subn(pattern, replacement, html, count=1)
    if count == 0:
        raise RuntimeError("未找到 'const DATA = { ... };' 块，请检查 HTML 文件")

    # 替换密码哈希占位符
    pwd_hash = hashlib.sha256(DASHBOARD_PASSWORD.encode("utf-8")).hexdigest()
    new_html = new_html.replace("__NPI_PWD_HASH__", pwd_hash)

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(new_html)


def main():
    print(f"读取 Excel: {EXCEL_PATH}")
    records = read_excel()
    print(f"共读取 {len(records)} 条记录")

    # 保存 JSON 副本
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump({"records": records, "buildTime": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")}, f, ensure_ascii=False, indent=2)
    print(f"JSON 已保存: {JSON_PATH}")

    # 嵌入 HTML
    inject_html(records)
    print(f"HTML 已更新: {HTML_PATH}")
    print("完成！双击打开 HTML 即可查看 Dashboard。")


if __name__ == "__main__":
    main()
